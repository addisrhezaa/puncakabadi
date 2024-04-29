from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import Http404, HttpResponse
from django.urls import reverse
from . import models
from django.db.models import Sum, Max
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import calendar

# Create your views here.


def gethargabahanbaku(
    listtanggal, hargamasukobj, hargakeluarobj, hargaawal, jumlahawal
):
    totalharga = hargaawal * jumlahawal
    for j in listtanggal:
        jumlahmasukperhari = 0
        hargamasuktotalperhari = 0
        hargamasuksatuanperhari = 0
        jumlahkeluarperhari = 0
        hargakeluartotalperhari = 0
        hargakeluarsatuanperhari = 0
        jumlahmasukperhari = 0

        sjpobj = hargamasukobj.filter(NoSuratJalan__Tanggal=j)
        if sjpobj.exists():
            for k in sjpobj:
                hargamasuktotalperhari += k.Harga * k.Jumlah
                jumlahmasukperhari += k.Jumlah
                hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
        else:
            hargamasuktotalperhari = 0
            jumlahmasukperhari = 0
            hargamasuksatuanperhari = 0
        # print(j)
        transaksigudangobj = hargakeluarobj.filter(tanggal=j)
        if transaksigudangobj.exists():
            for k in transaksigudangobj:
                jumlahkeluarperhari += k.jumlah
                hargakeluartotalperhari += k.jumlah * hargaawal
            hargakeluarsatuanperhari += hargakeluartotalperhari / jumlahkeluarperhari
        else:
            hargakeluartotalperhari = 0
            hargakeluarsatuanperhari = 0
            jumlahkeluarperhari = 0

        # print(
        #     "Tanggal : ",
        # )
        # print("Sisa Stok Hari Sebelumnya : ", jumlahawal)
        # print("harga awal Hari Sebelumnya :", hargaawal)
        # print("harga total Hari Sebelumnya :", totalharga)
        # print("Jumlah Masuk : ", jumlahmasukperhari)
        # print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
        # print("Harga Total Masuk : ", hargamasuktotalperhari)
        # print("Jumlah Keluar : ", jumlahkeluarperhari)
        # print("Harga Keluar : ", hargakeluarsatuanperhari)
        # print(
        #     "Harga Total Keluar : ",
        #     hargakeluarsatuanperhari * jumlahkeluarperhari,
        # )
        jumlahawal += jumlahmasukperhari - jumlahkeluarperhari
        totalharga += hargamasuktotalperhari - hargakeluartotalperhari
        try:
            hargaawal = totalharga / jumlahawal
        except ZeroDivisionError:
            hargaawal = 0

        # print("Sisa Stok Hari Ini : ", jumlahawal)
        # print("harga awal Hari Ini :", hargaawal)
        # print("harga total Hari Ini :", totalharga, "\n")
    return hargaawal, jumlahawal, totalharga


def dashboard(request):
    return render(request, "ppic/dashboard.html")


def laporanbarangjadi(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanstokfg.html")
    else:
        # Rumus = Saldo awal periode sampai tanggal akhir - Keluar awal periode sampai tanggal akhir
        tanggal_mulai = request.GET["tanggalawal"]
        tanggal_akhir = request.GET["tanggalakhir"]
        data = models.Artikel.objects.all()
        grandtotal = 0
        for i in data:
            mutasifilterobj = models.TransaksiProduksi.objects.filter(
                KodeArtikel=i.id
                # Tanggal__range=(tanggal_mulai, tanggal_akhir),
            )
            print(mutasifilterobj)
            saldomutasimasuktanggalakhir = mutasifilterobj.filter(
                Lokasi=1, Tanggal__lte=(tanggal_akhir), Jenis="Mutasi"
            )
            saldomutasikeluartanggalakhir = mutasifilterobj.filter(
                Lokasi=2, Tanggal__lte=(tanggal_akhir), Jenis="Mutasi"
            )
            jumlahmasuk = 0
            # jUMLAH kELUAR BELUM SYNC DENGAN MUTASI SPPB
            jumlahkeluar = 0
            for j in saldomutasimasuktanggalakhir:
                jumlahmasuk += j.Jumlah
            for K in saldomutasikeluartanggalakhir:
                jumlahkeluar += K.Jumlah
            i.Jumlahakumulasi = jumlahmasuk - jumlahkeluar
            print(jumlahmasuk)
            print(jumlahkeluar)
            # Nilai FG --> penyusun artikel * konversi * harga di akumulasikan semua penyusun
            penyusunfilterobj = models.Penyusun.objects.filter(KodeArtikel=i.id)

            nilaiFG = 0
            for penyusunobj in penyusunfilterobj:
                nilaiFG += gethargafg(penyusunobj)
            i.HargaFG = nilaiFG
            i.NilaiTotal = nilaiFG * i.Jumlahakumulasi
            grandtotal += i.NilaiTotal

        return render(
            request,
            "ppic/views_laporanstokfg.html",
            {
                "data": data,
                "tanggalawal": tanggal_mulai,
                "tanggalakhir": tanggal_akhir,
                "grandtotal": grandtotal,
            },
        )


def excel_laporanbarangmasuk(request):
    if request.method == "POST":
        tanggalawal = request.POST["tanggalawal"]
        tanggalakhir = request.POST["tanggalakhir"]
        dataspk = models.SuratJalanPembelian.objects.filter(
            Tanggal__range=(tanggalawal, tanggalakhir)
        ).order_by("Tanggal")
        # print(dataspk)
        listdetailsjp = []
        grandtotal = 0
        for i in dataspk:
            detailsjpembelianobj = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan=i.NoSuratJalan
            )
            for j in detailsjpembelianobj:
                j.supplier = i.supplier
                j.totalharga = j.Jumlah * j.Harga
                grandtotal += j.totalharga
                listdetailsjp.append(j)

        # print(listdetailsjp)
        listnomor = []
        listsjp = []
        listsupplier = []
        listkodeproduk = []
        listnamabarang = []
        listsatuan = []
        listqty = []
        listhargasatuan = []
        listhargatotal = []
        for no, i in enumerate(listdetailsjp):
            listnomor.append(no)
            listsjp.append(i.NoSuratJalan.NoSuratJalan)
            listsupplier.append(i.supplier)
            listkodeproduk.append(i.KodeProduk.KodeProduk)
            listnamabarang.append(i.KodeProduk.NamaProduk)
            listsatuan.append(i.KodeProduk.unit)
            listqty.append(i.Jumlah)
            listhargasatuan.append(i.Harga)
            listhargatotal.append(i.totalharga)
        tabel = {
            "No": listnomor,
            "SJ Pembelian": listsjp,
            "Supplier": listsupplier,
            "Kode Stok": listkodeproduk,
            "Nama Barang": listnamabarang,
            "Satuan": listsatuan,
            "Kuantitas": listqty,
            "Harga Satuan": listhargasatuan,
            "Harga Total": listhargatotal,
        }
        df = pd.DataFrame(tabel)
        print(df)

        excel_file = BytesIO()
        # df.to_excel('export data.xlsx',index=False)
        wb = Workbook()
        ws = wb.active
        headers = list(df.columns)
        ws.append(headers)
        for r_idx, row in enumerate(df.itertuples(), start=1):
            for c_idx, value in enumerate(row[1:], start=1):
                ws.cell(row=r_idx + 1, column=c_idx, value=value)

        start_row = (
            len(df) + 1
        )  # Ganti angka 2 dengan jumlah baris tambahan sebelum merge
        end_row = start_row  # Ganti angka 4 dengan jumlah baris yang ingin digabungkan
        totalhargacell = f"I{end_row}"
        labeltotalhargacell = f"H{end_row}"
        ws[totalhargacell] = grandtotal
        ws[labeltotalhargacell] = "Total Harga"
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        file_name = f"Laporan_{tanggalawal}-{tanggalakhir}.xlsx"
        response = HttpResponse(
            excel_buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        # with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
        #     df.to_excel(writer, index=False)
        # excel_file.seek(0)
        # response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = f'attachment; filename="Laporan{tanggalawal}-{tanggalakhir}.xlsx"'

        return response


def laporanbarangmasuk(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanbarangmasuk.html")
    else:
        tanggalawal = request.GET["tanggalawal"]
        tanggalakhir = request.GET["tanggalakhir"]

        dataspk = models.SuratJalanPembelian.objects.filter(
            Tanggal__range=(tanggalawal, tanggalakhir)
        ).order_by("Tanggal")
        print(dataspk)
        listdetailsjp = []
        grandtotal = 0
        for i in dataspk:
            detailsjpembelianobj = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan=i.NoSuratJalan
            )
            for j in detailsjpembelianobj:
                j.supplier = i.supplier
                j.totalharga = j.Jumlah * j.Harga
                grandtotal += j.totalharga
                listdetailsjp.append(j)

        print(listdetailsjp)
        return render(
            request,
            "ppic/views_laporanbarangmasuk.html",
            {
                "data": listdetailsjp,
                "tanggalawal": tanggalawal,
                "tanggalakhir": tanggalakhir,
                "grandtotal": grandtotal,
            },
        )


def gethargafg(penyusunobj):
    # Mengambil data Konversi master
    konversiobj = models.KonversiMaster.objects.get(
        KodePenyusun=penyusunobj.IDKodePenyusun
    )
    # Mengambil kuantitas dan ditambahkan 2.5%
    konversialowance = konversiobj.Kuantitas + (konversiobj.Kuantitas * 0.025)
    # Mengambil data Detail surrat jalan pembelian untuk kode produk sesuai dengan penyusun obj (1 penyusun 1 produk 1 artkel)
    detailsjpembelian = models.DetailSuratJalanPembelian.objects.filter(
        KodeProduk=penyusunobj.KodeProduk
    )
    hargatotalkodeproduk = 0
    jumlahtotalkodeproduk = 0
    for m in detailsjpembelian:
        hargatotalkodeproduk += m.Harga * m.Jumlah
        jumlahtotalkodeproduk += m.Jumlah
    # print("ini jumlah harga total ", hargatotalkodeproduk)
    try:
        rataratahargakodeproduk = hargatotalkodeproduk / jumlahtotalkodeproduk
    except ZeroDivisionError:
        rataratahargakodeproduk = 0
    # print("selesai")
    # print(rataratahargakodeproduk)
    nilaifgperkodeproduk = rataratahargakodeproduk * konversialowance
    # print("Harga Konversi : ", nilaifgperkodeproduk)
    return nilaifgperkodeproduk


def laporanbarangkeluar(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanbarangkeluar.html")
    else:
        tanggalawal = request.GET["tanggalawal"]
        tanggalakhir = request.GET["tanggalakhir"]
        data = models.SPPB.objects.filter(
            Tanggal__range=(tanggalawal, tanggalakhir)
        ).order_by("Tanggal")
        print("ini data", data)
        listharga = []
        listdata = []
        listkodeartikel = []
        listjumlah = []
        listhargafg = []
        listnilaitotal = []
        datakirim = []
        if not data.exists():
            messages.warning(
                request, "Data SPPB tidak ditemukan pada rentang tanggal tersebut"
            )
            return redirect("laporanbarangkeluar")
        for i in data:
            detailsppb = models.DetailSPPB.objects.filter(NoSPPB=i.id)
            a = detailsppb.values("DetailSPK__KodeArtikel").annotate(
                total_jumlah=Sum("Jumlah")
            )
            print("nilai A", a)
            for j in a:
                print(j["DetailSPK__KodeArtikel"])
                kodeartikel = j["DetailSPK__KodeArtikel"]
                penyusunfilterobj = models.Penyusun.objects.filter(
                    KodeArtikel=kodeartikel
                )
                if kodeartikel in listkodeartikel:
                    index = listkodeartikel.index(kodeartikel)
                    jumlah = listjumlah[index] + j["total_jumlah"]
                    listjumlah[index] = jumlah
                    listnilaitotal[index] = jumlah * listhargafg[index]
                else:
                    listkodeartikel.append(kodeartikel)
                    jumlah = j["total_jumlah"]
                    listjumlah.append(jumlah)

                    nilaiFG = 0
                    for penyusunobj in penyusunfilterobj:

                        nilaiFG += gethargafg(penyusunobj)
                    listhargafg.append(nilaiFG)
                    listnilaitotal.append(nilaiFG * jumlah)

            listdata.append(a)
        # print(listdata)
        grandtotal = sum(listnilaitotal)
        print("listkodeartikel", listkodeartikel)
        print("listjumlah", listjumlah)
        print("listnilaitotal", listnilaitotal)
        print("listhargafg", listhargafg)

        for kode_artikel, jumlah, nilai_total, harga_fg in zip(
            listkodeartikel, listjumlah, listnilaitotal, listhargafg
        ):
            artikel = models.Artikel.objects.get(id=kode_artikel)
            datakirim.append(
                {
                    "kode_artikel": artikel,
                    "jumlah": jumlah,
                    "nilai_total": nilai_total,
                    "harga_fg": harga_fg,
                }
            )

        return render(
            request,
            "ppic/views_laporanbarangkeluar.html",
            {
                "tanggalawal": tanggalawal,
                "tanggalakhir": tanggalakhir,
                "data": datakirim,
                "grandtotal": grandtotal,
            },
        )


def laporanpersediaanbarang(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanpersediaan.html")
    else:
        tanggal_mulai = request.GET["tanggalawal"]
        tanggal_akhir = request.GET["tanggalakhir"]
        tanggal_obj = datetime.strptime(tanggal_akhir, "%Y-%m-%d").date()

        # Ambil total harga barang keluar dulu
        data = models.SPPB.objects.filter(
            Tanggal__range=(tanggal_mulai, tanggal_akhir)
        ).order_by("Tanggal")
        if not data.exists():
            messages.warning(
                request, "Data SPPB Tidak ditemukan pada rentang tanggal tersebut"
            )
            return redirect("laporanpersediaanbarang")
        listharga = []
        for i in data:
            detailsppb = models.DetailSPPB.objects.filter(NoSPPB=i.id)
            a = detailsppb.values("DetailSPK__KodeArtikel").annotate(
                total_jumlah=Sum("Jumlah")
            )
            for j in a:
                penyusunfilterobj = models.Penyusun.objects.filter(
                    KodeArtikel=j["DetailSPK__KodeArtikel"]
                )
                nilaiFG = 0
                for penyusunobj in penyusunfilterobj:
                    nilaiFG += gethargafg(penyusunobj)
                j.update({"HargaFG": nilaiFG})
                j.update({"TotalNilai": nilaiFG * j["total_jumlah"]})
                j["DetailSPK__KodeArtikel"] = penyusunobj.KodeArtikel.KodeArtikel
                listharga.append(j["TotalNilai"])
        totalhargabarangkeluar = sum(listharga)

        # Ambil data Bahan masuk
        dataspk = models.SuratJalanPembelian.objects.filter(
            Tanggal__range=(tanggal_mulai, tanggal_akhir)
        ).order_by("Tanggal")
        listdetailsjp = []
        totalhargabarangmasuk = 0
        for i in dataspk:
            detailsjpembelianobj = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan=i.NoSuratJalan
            )
            for j in detailsjpembelianobj:
                j.supplier = i.supplier
                j.totalharga = j.Jumlah * j.Harga
                totalhargabarangmasuk += j.totalharga
                listdetailsjp.append(j)

        # Total Harga Stok Gudang
        dataartikel = models.Artikel.objects.all()
        totalhargabarangjadi = 0
        for i in dataartikel:
            mutasifilterobj = models.TransaksiProduksi.objects.filter(KodeArtikel=i.id)
            saldomutasimasuktanggalakhir = mutasifilterobj.filter(
                Lokasi=1, Tanggal__lte=(tanggal_akhir)
            )
            saldomutasikeluartanggalakhir = mutasifilterobj.filter(
                Lokasi=2, Tanggal__lte=(tanggal_akhir)
            )
            print(mutasifilterobj)
            jumlahmasuk = 0
            jumlahkeluar = 0
            for j in saldomutasimasuktanggalakhir:
                jumlahmasuk += j.Jumlah
            for k in saldomutasikeluartanggalakhir:
                jumlahkeluar += k.Jumlah

            i.Jumlahakumulasi = jumlahmasuk - jumlahkeluar
            penyusunfilterobj = models.Penyusun.objects.filter(KodeArtikel=i.id)

            nilaiFG = 0
            for penyusunobj in penyusunfilterobj:
                nilaiFG += gethargafg(penyusunobj)
            i.Harga = nilaiFG
            i.NilaiTotal = nilaiFG * i.Jumlahakumulasi
            totalhargabarangjadi += i.NilaiTotal

        # Saldo awal Belum dibuat
        Saldoawal = 0
        # Anggap dari
        # Ambil data semua bahan baku di wip
        databahanproduksi = models.Produk.objects.filter(KodeProduk__startswith="A")
        print(databahanproduksi)
        for i in databahanproduksi:
            print("ini i", i)
            datasaldoawaltahun = (
                models.SaldoAwalBahanBaku.objects.filter(
                    Tanggal__lte=tanggal_akhir, IDBahanBaku=i.KodeProduk
                )
                .order_by("-Tanggal")
                .first()
            )
            if (
                not datasaldoawaltahun
                or datasaldoawaltahun.Tanggal.year != tanggal_obj.year
            ):
                print("data tidak ada")
                hargasaldoawalbahanbaku = 0
            else:
                print(datasaldoawaltahun)
                print("data ada")
                hargasaldoawalbahanbaku = (
                    datasaldoawaltahun.Harga * datasaldoawaltahun.Jumlah
                )

            Saldoawal += hargasaldoawalbahanbaku

        #
        saldototal = Saldoawal + totalhargabarangmasuk - totalhargabarangkeluar
        saldowip = saldototal - totalhargabarangjadi

        return render(
            request,
            "ppic/views_laporanpersediaan.html",
            {
                "tanggalawal": tanggal_mulai,
                "tanggalakhir": tanggal_akhir,
                "data": a,
                "barangkeluar": round(totalhargabarangkeluar, 2),
                "barangmasuk": round(totalhargabarangmasuk, 2),
                "barangfg": round(totalhargabarangjadi, 2),
                "saldoawal": round(Saldoawal, 2),
                "saldototal": round(saldototal, 2),
                "saldowip": round(saldowip, 2),
            },
        )


"""
Revisi 4/21/2024
1. Co PO
2. Transaction Log
3. Perhitungan Laporan
"""


def viewconfirmationorder(request):
    data = models.confirmationorder.objects.all()
    for i in data:
        detailcopo = models.detailconfirmationorder.objects.filter(
            confirmationorder=i.id
        )
        i.detailcopo = detailcopo
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    if len(request.GET) == 0:
        return render(request, "ppic/views_confirmationorder.html", {"data": data})


def tambahconfirmationorder(request):
    dataartikel = models.Artikel.objects.all()
    if request.method == "GET":
        return render(request, "ppic/add_co.html", {"dataartikel": dataartikel})
    else:
        print(request.POST)
        tanggaladd = request.POST["tanggal"]
        nomorco = request.POST["nomorco"]
        kepada = request.POST["kepada"]
        perihal = request.POST["perihal"]
        artikel = request.POST.getlist("artikel[]")
        kuantitas = request.POST.getlist("kuantitas[]")
        harga = request.POST.getlist("harga[]")
        deskripsi = request.POST.getlist("deskripsi[]")
        # print(tanggaladd)
        # print(nomorco)
        # print(kepada)
        # print(perihal)
        # print(artikel)
        # print(kuantitas)
        # print(harga)
        # print(deskripsi)

        confirmationorderobj = models.confirmationorder(
            NoCO=nomorco, kepada=kepada, perihal=perihal, tanggal=tanggaladd
        )
        confirmationorderobj.save()
        print(confirmationorderobj.id)
        for artikel, kuantitas, harga, deskripsi in zip(
            artikel, kuantitas, harga, deskripsi
        ):
            # print(artikel, kuantitas, harga, deskripsi)
            detailconfirmationobj = models.detailconfirmationorder(
                confirmationorder=confirmationorderobj,
                Artikel=models.Artikel.objects.get(KodeArtikel=artikel),
                Harga=harga,
                kuantitas=kuantitas,
                deskripsi=deskripsi,
            )
            print(dir(detailconfirmationobj))
            detailconfirmationobj.save()
        return redirect("confirmationorder")


def detailco(request, id):
    data = models.confirmationorder.objects.get(id=id)
    detailcopo = models.detailconfirmationorder.objects.filter(
        confirmationorder=data.id
    )
    data.detailcopo = detailcopo
    data.tanggal = data.tanggal.strftime("%Y-%m-%d")

    return render(request, "ppic/detailco.html", {"dataco": data})


def updateco(request, id):
    data = models.confirmationorder.objects.get(id=id)
    detailcopo = models.detailconfirmationorder.objects.filter(
        confirmationorder=data.id
    )
    data.detailcopo = detailcopo
    data.tanggal = data.tanggal.strftime("%Y-%m-%d")
    print(len(data.detailcopo))
    if request.method == "GET":
        return render(request, "ppic/updateco.html", {"dataco": data})
    else:
        print(request.POST)
        tanggaladd = request.POST["tanggal"]
        nomorco = request.POST["nomorco"]
        kepada = request.POST["kepada"]
        perihal = request.POST["perihal"]
        artikel = request.POST.getlist("artikel[]")
        kuantitas = request.POST.getlist("kuantitas[]")
        harga = request.POST.getlist("harga[]")
        deskripsi = request.POST.getlist("deskripsi[]")
        listid = request.POST.getlist("id[]")
        data.tanggal = tanggaladd
        data.nomorco = nomorco
        data.kepada = kepada
        data.perihal = perihal

        data.save()

        for listid, artikel, kuantitas, harga, deskripsi in zip(
            listid, artikel, kuantitas, harga, deskripsi
        ):
            # print(artikel, kuantitas, harga, deskripsi)
            if listid == "":
                detailconfirmationobj = models.detailconfirmationorder(
                    confirmationorder=data,
                    Artikel=models.Artikel.objects.get(KodeArtikel=artikel),
                    Harga=harga,
                    kuantitas=kuantitas,
                    deskripsi=deskripsi,
                )
            else:
                detailconfirmationobj = models.detailconfirmationorder.objects.get(
                    id=listid
                )
                detailconfirmationobj.confirmationorder = data
                detailconfirmationobj.Artikel = models.Artikel.objects.get(
                    KodeArtikel=artikel
                )
                detailconfirmationobj.Harga = harga
                detailconfirmationobj.kuantitas = kuantitas
                detailconfirmationobj.deskripsi = deskripsi

            detailconfirmationobj.save()

        return redirect("confirmationorder")


def deletedetailco(request, id):
    data = models.detailconfirmationorder.objects.get(id=id)
    idco = data.confirmationorder.id
    data.delete()
    return redirect("updateco", id=idco)


def deleteco(request, id):
    data = models.confirmationorder.objects.get(id=id)
    data.delete()
    return redirect("confirmationorder")


def newlaporanpersediaan(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_newlaporanpersediaan.html")
    else:
        """Initiation"""
        tanggalmulai = request.GET["tanggalawal"]
        tanggalakhir = request.GET["tanggalakhir"]
        tanggal_obj = datetime.strptime(tanggalakhir, "%Y-%m-%d")
        tahun = tanggal_obj.year
        awaltahun = datetime(tahun, 1, 1)
        tanggal_obj.date()

        data = models.SPPB.objects.filter(
            Tanggal__range=(tanggalmulai, tanggalakhir)
        ).order_by("Tanggal")
        if not data.exists():
            messages.warning(
                request, "Data SPPB Tidak ditemukan pada rentang tanggal tersebut"
            )
        """End Initiation"""

        """
        SECTION SJP
        sudah clear.
        SJP Hanya mempertimbangkan transaksi di range tanggal tersebut. Tidak ada cakupan untuk perhitungan yang lain
        """

        datasjp = models.SuratJalanPembelian.objects.filter(
            Tanggal__range=(tanggalmulai, tanggalakhir)
        ).order_by("Tanggal")
        listdetailsjp = []
        totalhargabarangmasuk = 0
        for i in datasjp:
            detailsjpembelianobj = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan=i.NoSuratJalan
            )
            for j in detailsjpembelianobj:
                j.supplier = i.supplier
                j.totalharga = j.Jumlah * j.Harga
                totalhargabarangmasuk += j.totalharga
                listdetailsjp.append(j)
        # print("total barang masuk SJP : ", totalhargabarangmasuk)
        # Output : total barang masuk SJP :  45581000.0
        """ END SECTION SJP """

        """
        TOTAL HARGA STOK
        Belum FIx untuk perhitungan Harga.
        Bisa di cek dulu menggunakan harga terakhir di bulan tersebut. 
        apabila ada harganya maka bisa menggunakan harga tersebut. Apabila tidak ada maka menggunakan harga terbesar tanggal terakhir
        Apabila cek kondisi 0 maka muncul warning dan memilih harga terakhir
        """
        bahanbaku = models.Produk.objects.all()
        # bahanbaku = models.Produk.objects.filter(KodeProduk="coba-001")
        listhargabahanbaku = {}
        stokakhirbahanbakutiapbulan = {}
        hargaakhirbulanperproduk = {}
        """ PERHITUNGAN HARGA BARU """
        for i in bahanbaku:
            hargamasukobj = models.DetailSuratJalanPembelian.objects.filter(
                KodeProduk=i, NoSuratJalan__Tanggal__gte=awaltahun
            )
            hargakeluarobj = models.TransaksiGudang.objects.filter(
                KodeProduk=i, tanggal__gte=awaltahun
            )
            tanggalhargamasukobj = (
                hargamasukobj.filter(NoSuratJalan__Tanggal__lte=tanggalakhir)
                .values_list("NoSuratJalan__Tanggal", flat=True)
                .distinct()
            )
            tanggalhargakeluarobj = (
                hargakeluarobj.filter(tanggal__lte=tanggalakhir)
                .values_list("tanggal", flat=True)
                .distinct()
            )
            listtanggal = sorted(
                list(set(tanggalhargamasukobj.union(tanggalhargakeluarobj)))
            )

            try:
                saldoawalobj = models.SaldoAwalBahanBaku.objects.get(
                    IDBahanBaku=i, Tanggal__gte=awaltahun
                )
                hargaawal = saldoawalobj.Harga
                jumlahawal = saldoawalobj.Jumlah

            except models.SaldoAwalBahanBaku.DoesNotExist:
                hargaawal = 0
                jumlahawal = 0
            # Menghitung harga tanggal akhir
            data = gethargabahanbaku(
                listtanggal,
                hargamasukobj,
                hargakeluarobj,
                hargaawal,
                jumlahawal,
            )
            listhargabahanbaku[i] = data[0]

            """
            V2 List harga bahan baku awal dipisah berdasarkan tanggal
            1. Ambil dulu list tanggal akhir tiap bulan
            2. cek harga tiap akhir bulan 
            A. Filter dari GTE awal taun dan LTE akhir bulan untuk SJP dan Transaksi Gudang
            """
            """ 1. Ambil List Akhir bulan"""
            # print(calendar.monthrange(tahun, 1))
            last_days = []
            for month in range(1, 13):
                last_day = calendar.monthrange(tahun, month)[1]
                last_days.append(date(tahun, month, last_day))

            """ End list akhir bulan"""

            """ 2. Cek harga akhir Bulan"""
            hargaakhirbulan = {}
            maxtanggal = hargamasukobj.aggregate(Max("NoSuratJalan__Tanggal"))[
                "NoSuratJalan__Tanggal__max"
            ].month
            # print("Maksimal data", hargamasukobj)
            # print("Maksimal bulan", maxtanggal)
            # print(last_days[:maxtanggal])
            totalhargabahanbakugudangperbulan = 0
            for j, k in enumerate(last_days[:maxtanggal]):
                suratjalanpembelianakhirbulanobj = hargamasukobj.filter(
                    NoSuratJalan__Tanggal__lte=k
                )
                transaksigudangakhirbulanobj = hargakeluarobj.filter(tanggal__lte=k)
                tanggalsuratjalanpembelianakhirbulanobj = (
                    suratjalanpembelianakhirbulanobj.values_list(
                        "NoSuratJalan__Tanggal", flat=True
                    ).distinct()
                )
                tanggaltransaksigudangakhirbulanobj = (
                    transaksigudangakhirbulanobj.values_list(
                        "tanggal", flat=True
                    ).distinct()
                )
                tanggalkeluarmasukperbulan = sorted(
                    list(
                        set(
                            tanggalsuratjalanpembelianakhirbulanobj.union(
                                tanggaltransaksigudangakhirbulanobj
                            )
                        )
                    )
                )
                datahargaperbulan = gethargabahanbaku(
                    tanggalkeluarmasukperbulan,
                    suratjalanpembelianakhirbulanobj,
                    transaksigudangakhirbulanobj,
                    hargaawal,
                    jumlahawal,
                )
                # print(j, k, datahargaperbulan)
                hargaakhirbulan[j] = {
                    "hargasatuan": datahargaperbulan[0],
                    "jumlah": datahargaperbulan[1],
                    "hargatotal": datahargaperbulan[2],
                }
                totalhargabahanbakugudangperbulan += datahargaperbulan[2]
            hargaakhirbulanperproduk[i] = {
                "data": hargaakhirbulan,
                "total": totalhargabahanbakugudangperbulan,
            }
            # print("data surat jalan pembelian : ", suratjalanpembelianakhirbulanobj)
            # print("data Transaksi Gudang : ", transaksigudangakhirbulanobj)

            """ End harga akhir bulan"""
            """ endsection V2 """

        #     print("\n\nDONES\n\n")
        # print(listhargabahanbaku)
        # print(listhargabahanbakusebelum)
        """ END SECTION PERHITUNGAN HARGA BARU """

        print("Data Rekap Harga Perbulan : ", hargaakhirbulanperproduk)
        # print(asdasdas)
        totalhargabarangjadi = 0
        """
        Data Models harga FG perbulan
        {
            Kode Artikel : [
            { bulan0 :
                {
            item1 : 9000
            item2 : 8000
            }
            HargaFG : 17000
            }
            ]
        }
        """
        dataartikel = models.Artikel.objects.all()
        datahargafgartikel = {}
        for artikel in dataartikel:
            hargafg = 0
            dataperbulan = {}
            """
            Models perbulan
            {}
            """
            modelsperbulan = {}
            for index, hari in enumerate(last_days[: tanggal_obj.month]):
                # Mengambil data terakhir versi penyusun tiap bulannya
                versiterakhirperbulan = (
                    models.Penyusun.objects.filter(KodeArtikel=artikel, versi__lte=hari)
                    .values_list("versi", flat=True)
                    .distinct()
                    .order_by("versi")
                    .last()
                )
                # SEMENTARA PAKAI .LAST()
                penyusunversiterpilih = models.Penyusun.objects.filter(
                    KodeArtikel=artikel, versi=versiterakhirperbulan
                )
                datapenyusun = {}
                for penyusun in penyusunversiterpilih:
                    dummy = {}
                    hargapenyusun = hargaakhirbulanperproduk[penyusun.KodeProduk][
                        "data"
                    ][index]["hargasatuan"]
                    kuantitas = models.KonversiMaster.objects.get(
                        KodePenyusun=penyusun
                    ).Kuantitas
                    hargabahanbakufg = hargapenyusun * kuantitas
                    hargafg += hargabahanbakufg
                    dummy["totalharga"] = hargabahanbakufg
                    dummy["kuantitas"] = kuantitas
                    dummy["harga"] = hargapenyusun
                    datapenyusun[penyusun.KodeProduk] = dummy

                dummy2 = {}
                dummy2["penyusun"] = datapenyusun
                dummy2["hargafg"] = hargafg

                dataperbulan[index] = dummy2
            datahargafgartikel[artikel] = dataperbulan

        # print("versi terakhir ", datahargafgartikel)
        rekapkeluarperbulan = {}
        rekapmutasiperbulan = {}
        sisaperbulan = {}
        bahangudangperbulan = {}
        datarekapbarangmasukperbulan = {}

        for index, hari in enumerate(last_days[: tanggal_obj.month]):
            """Section SJP --> Mencari barnag masu tiap bulan"""
            nilaibahanbakumasuk = 0
            dataartikelmasuk = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__lte=hari,
                NoSuratJalan__Tanggal__gte=awaltahun,
            )
            dummy = {}
            for item in dataartikelmasuk:
                biaya = item.Harga * item.Jumlah
                nilaibahanbakumasuk += biaya
                dummy[item] = {
                    "biaya": biaya,
                    "jumlah": item.Jumlah,
                    "harga": item.Harga,
                    "kodeproduk": item.KodeProduk,
                }
            datarekapbarangmasukperbulan[index] = {
                "data": dummy,
                "biayatotal": nilaibahanbakumasuk,
            }
            """Section SPPB --> mencari rekap barnag keluar tiap bulan"""
            if index > 0:
                datadetailsppb = models.DetailSPPB.objects.filter(
                    NoSPPB__Tanggal__lte=hari, NoSPPB__Tanggal__gt=last_days[index - 1]
                )
                datatransaksiproduksi = models.TransaksiProduksi.objects.filter(
                    Tanggal__lte=hari, Tanggal__gt=last_days[index - 1], Jenis="Mutasi"
                )
            else:
                datadetailsppb = models.DetailSPPB.objects.filter(
                    NoSPPB__Tanggal__lte=hari, NoSPPB__Tanggal__gte=awaltahun
                )
                datatransaksiproduksi = models.TransaksiProduksi.objects.filter(
                    Tanggal__lte=hari, Tanggal__gt=awaltahun, Jenis="Mutasi"
                )
            print("Data detail SPPB\n", index, datadetailsppb)
            jumlahkumulatifbiayaperbulan = 0
            if datadetailsppb.exists():
                jumlahartikelkeluarperbulan = datadetailsppb.values(
                    "DetailSPK__KodeArtikel"
                ).annotate(total=Sum("Jumlah"))
                # print(jumlahartikelkeluarperbulan)
                dummy = {}
                for artikel in jumlahartikelkeluarperbulan:
                    dataartikel = models.Artikel.objects.get(
                        id=artikel["DetailSPK__KodeArtikel"]
                    )

                    jumlah = artikel["total"]
                    totalbiaya = (
                        jumlah * datahargafgartikel[dataartikel][index]["hargafg"]
                    )
                    # print(dataartikel, jumlah)
                    # print(datahargafgartikel[dataartikel][index]["hargafg"], totalbiaya)
                    dummy[dataartikel] = {
                        "jumlah": jumlah,
                        "hargafg": datahargafgartikel[dataartikel][index]["hargafg"],
                        "biaya": totalbiaya,
                    }
                    jumlahkumulatifbiayaperbulan += totalbiaya
            else:
                dummy = 0
            rekapkeluarperbulan[index] = {
                "data": dummy,
                "jumlah": jumlahkumulatifbiayaperbulan,
            }
            dummy = {}
            """SECTION Transaksi Produksi --> Untuk mencari data mutasi barang jadi"""
            if datatransaksiproduksi.exists():
                jumlahartikelmutasiperbulan = datatransaksiproduksi.values(
                    "KodeArtikel"
                ).annotate(total=Sum("Jumlah"))
                for artikel in jumlahartikelmutasiperbulan:
                    dataartikel = models.Artikel.objects.get(id=artikel["KodeArtikel"])
                    jumlah = artikel["total"]
                    totalbiaya = (
                        jumlah * datahargafgartikel[dataartikel][index]["hargafg"]
                    )
                    dummy[dataartikel] = {
                        "jumlah": jumlah,
                        "hargafg": datahargafgartikel[dataartikel][index]["hargafg"],
                        "biaya": totalbiaya,
                    }
            else:
                dummy = 0

            rekapmutasiperbulan[index] = dummy
            dataartikel = models.Artikel.objects.all()
            dummy = {}
            totalbiayasisafg = 0
            for artikel in dataartikel:
                if index == 0:
                    dummy[artikel] = {"jumlah": 0, "biaya": 0, "hargafg": 0}
                else:
                    dummy[artikel] = {
                        "jumlah": sisaperbulan[index - 1]["data"][artikel]["jumlah"]
                    }
                # ambil jumlah artikel keluar terkait
                if (
                    datadetailsppb.exists()
                    and artikel in rekapkeluarperbulan[index]["data"]
                ):
                    jumlahartikelkeluar = rekapkeluarperbulan[index]["data"][artikel][
                        "jumlah"
                    ]
                else:
                    jumlahartikelkeluar = 0

                if (
                    datatransaksiproduksi.exists()
                    and artikel in rekapmutasiperbulan[index]
                ):
                    jumlahartikelmasuk = rekapmutasiperbulan[index][artikel]["jumlah"]
                else:
                    jumlahartikelmasuk = 0
                print(artikel, jumlahartikelkeluar, jumlahartikelmasuk)

                # print(dummy)
                dummy[artikel]["jumlah"] += jumlahartikelmasuk - jumlahartikelkeluar
                dummy[artikel]["hargafg"] = datahargafgartikel[artikel][index][
                    "hargafg"
                ]
                dummy[artikel]["biaya"] = (
                    dummy[artikel]["jumlah"]
                    * datahargafgartikel[artikel][index]["hargafg"]
                )
                # print(dummy)
                totalbiayasisafg += dummy[artikel]["biaya"]
            sisaperbulan[index] = {"data": dummy, "total": totalbiayasisafg}

            """SECTION Stock Gudang"""
            bahanbaku = models.Produk.objects.all()
            # bahanbaku = models.Produk.objects.filter(KodeProduk="A-001-06")
            dummy = {}
            rekaphargabahanbakugudangperbulan = 0
            for produk in bahanbaku:
                saldoawalobj = models.SaldoAwalBahanBaku.objects.filter(
                    IDBahanBaku=produk, Tanggal__gte=awaltahun
                )
                if saldoawalobj.exists():
                    saldoawalobj = saldoawalobj.first()
                    totalbiayaawal = saldoawalobj.Harga * saldoawalobj.Jumlah
                    hargasatuanawal = saldoawalobj.Harga
                    jumlahawal = saldoawalobj.Jumlah
                else:
                    totalbiayaawal = 0
                    hargasatuanawal = 0
                    jumlahawal = 0

                barangmasukobj = models.DetailSuratJalanPembelian.objects.filter(
                    KodeProduk=produk,
                    NoSuratJalan__Tanggal__gte=awaltahun,
                    NoSuratJalan__Tanggal__lte=hari,
                )
                barangkeluarobj = models.TransaksiGudang.objects.filter(
                    KodeProduk=produk, tanggal__gte=awaltahun, tanggal__lte=hari
                )

                tanggalbarangmasukobj = barangmasukobj.values_list(
                    "NoSuratJalan__Tanggal", flat=True
                ).distinct()
                tanggalbarangkeluar = barangkeluarobj.values_list(
                    "tanggal", flat=True
                ).distinct()
                listtanggal = sorted(
                    list(set(tanggalbarangmasukobj.union(tanggalbarangkeluar)))
                )
                print(index, produk, listtanggal)
                for tanggal in listtanggal:
                    jumlahmasukperhari = 0
                    hargamasuktotalperhari = 0
                    hargamasuksatuanperhari = 0
                    jumlahkeluarperhari = 0
                    hargakeluartotalperhari = 0
                    hargakeluarsatuanperhari = 0
                    jumlahmasukperhari = 0

                    sjpobj = barangmasukobj.filter(NoSuratJalan__Tanggal=tanggal)
                    if sjpobj.exists():
                        for k in sjpobj:
                            hargamasuktotalperhari += k.Harga * k.Jumlah
                            jumlahmasukperhari += k.Jumlah
                            hargamasuksatuanperhari += (
                                hargamasuktotalperhari / jumlahmasukperhari
                            )
                    else:
                        hargamasuktotalperhari = 0
                        jumlahmasukperhari = 0
                        hargamasuksatuanperhari = 0

                    transaksigudangobj = barangkeluarobj.filter(tanggal=tanggal)
                    if transaksigudangobj.exists():
                        for k in transaksigudangobj:
                            jumlahkeluarperhari += k.jumlah
                            hargakeluartotalperhari += k.jumlah * hargasatuanawal
                        hargakeluarsatuanperhari += (
                            hargakeluartotalperhari / jumlahkeluarperhari
                        )
                    else:
                        hargakeluartotalperhari = 0
                        hargakeluarsatuanperhari = 0
                        jumlahkeluarperhari = 0

                    # print(produk)
                    # print("Tanggal : ", tanggal)
                    # print("Sisa Stok Hari Sebelumnya : ", jumlahawal)
                    # print("harga awal Hari Sebelumnya :", hargasatuanawal)
                    # print("harga total Hari Sebelumnya :", totalbiayaawal)
                    # print("Jumlah Masuk : ", jumlahmasukperhari)
                    # print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
                    # print("Harga Total Masuk : ", hargamasuktotalperhari)
                    # print("Jumlah Keluar : ", jumlahkeluarperhari)
                    # print("Harga Keluar : ", hargakeluarsatuanperhari)
                    # print(
                    #     "Harga Total Keluar : ",
                    #     hargakeluarsatuanperhari * jumlahkeluarperhari,
                    # )
                    jumlahawal += jumlahmasukperhari - jumlahkeluarperhari
                    totalbiayaawal += hargamasuktotalperhari - hargakeluartotalperhari
                    try:
                        hargasatuanawal = totalbiayaawal / jumlahawal
                    except ZeroDivisionError:
                        hargasatuanawal = 0

                    # print("Sisa Stok Hari Ini : ", jumlahawal)
                    # print("harga awal Hari Ini :", hargasatuanawal)
                    # print("harga total Hari Ini :", totalbiayaawal, "\n")
                dummy[produk] = {
                    "hargasatuan": hargasatuanawal,
                    "jumlah": jumlahawal,
                    "totalbiaya": totalbiayaawal,
                }
                rekaphargabahanbakugudangperbulan += totalbiayaawal

            bahangudangperbulan[index] = {
                "data": dummy,
                "total": rekaphargabahanbakugudangperbulan,
            }
        print("ini rekap keluar perbulan\n", rekapkeluarperbulan)
        print("ini rekap mutasi perbulan\n", rekapmutasiperbulan)
        print("ini rekap Sisa perbulan\n", sisaperbulan)
        print("ini rekap Bahan Gudang perbulan\n", bahangudangperbulan)
        # print(adasadaasd)

        """Rekapitulasi ke Models Akhir
        data models
        {
        Bulan1 : 
        {
        barangkeluar : {
        artikel1 :value,
        artikel2:value,
        dst
        },
        barangmasuk :{}
        bahanproduksi : {}
        baranggudang : {}
        }
        }
        """
        modelakhir = {}
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        for month in range(0, tanggal_obj.month):
            # Data Barang Keluar
            modelakhir[listbulan[month]] = {
                "barangkeluar": rekapkeluarperbulan[month]["jumlah"],
                "barangmasuk": datarekapbarangmasukperbulan[month]["biayatotal"],
                "detailbarangkeluar": rekapkeluarperbulan[month]["data"],
                "sisafg": sisaperbulan[month],
                "stockgudang": bahangudangperbulan[month],
            }
        return render(
            request,
            "ppic/views_newlaporanpersediaan.html",
            {"modeldata": modelakhir},
        )
